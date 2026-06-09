# -*- coding: utf-8 -*-
"""
独立爬取+导出脚本

不依赖 Flask，直接运行所有爬虫，将结果保存到数据库，并导出为 Excel 和 CSV。

用法：
    python crawl_export.py
"""

import os
import sys
import csv
from datetime import datetime

# 确保项目根目录在 path 中
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT_DIR)

from database import init_db, get_db  # noqa: E402


def _print_one(result):
    """格式化打印一个爬虫的运行结果"""
    if not isinstance(result, dict):
        return
    name = result.get('name', '?')
    status = result.get('status', '?')
    total = result.get('total', 0)
    inserted = result.get('inserted', 0)
    err = result.get('error') or ''
    if status == 'success':
        print(f"  [OK]   {name}: 共 {total} 条，新入库 {inserted} 条")
    else:
        print(f"  [FAIL] {name}: {err}")


def run_all_spiders():
    """运行所有爬虫，单个失败不影响其他"""
    # 延迟导入，避免依赖问题影响主流程
    try:
        from spiders.national import NATIONAL_SPIDERS, run_all_national
    except Exception as exc:  # noqa: BLE001
        NATIONAL_SPIDERS, run_all_national = [], None
        print(f"[警告] 国家级爬虫模块加载失败: {exc}")

    try:
        from spiders.sichuan import SICHUAN_SPIDERS, run_all_sichuan
    except Exception as exc:  # noqa: BLE001
        SICHUAN_SPIDERS, run_all_sichuan = [], None
        print(f"[警告] 四川省级爬虫模块加载失败: {exc}")

    try:
        from spiders.chengdu import CHENGDU_SPIDERS, run_all_chengdu
    except Exception as exc:  # noqa: BLE001
        CHENGDU_SPIDERS, run_all_chengdu = [], None
        print(f"[警告] 成都市级爬虫模块加载失败: {exc}")

    try:
        from spiders.gaoxintong import GaoXinTongSpider
    except Exception as exc:  # noqa: BLE001
        GaoXinTongSpider = None
        print(f"[警告] 高新通爬虫模块加载失败: {exc}")

    total_sites = (len(NATIONAL_SPIDERS) + len(SICHUAN_SPIDERS)
                   + len(CHENGDU_SPIDERS) + (1 if GaoXinTongSpider else 0))
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始爬取...")
    print(f"  共 {total_sites} 个网站\n")

    summary = {'total': 0, 'inserted': 0, 'success': 0, 'failed': 0}

    def _accumulate(results):
        if not results:
            return
        for r in results:
            _print_one(r)
            if isinstance(r, dict):
                summary['total'] += int(r.get('total') or 0)
                summary['inserted'] += int(r.get('inserted') or 0)
                if r.get('status') == 'success':
                    summary['success'] += 1
                else:
                    summary['failed'] += 1

    # 国家级
    if run_all_national:
        print("--- 国家级网站 ---")
        try:
            _accumulate(run_all_national())
        except Exception as exc:  # noqa: BLE001
            print(f"  [FAIL] 国家级整体: {exc}")

    # 省级
    if run_all_sichuan:
        print("\n--- 四川省级网站 ---")
        try:
            _accumulate(run_all_sichuan())
        except Exception as exc:  # noqa: BLE001
            print(f"  [FAIL] 省级整体: {exc}")

    # 市级
    if run_all_chengdu:
        print("\n--- 成都市级网站 ---")
        try:
            _accumulate(run_all_chengdu())
        except Exception as exc:  # noqa: BLE001
            print(f"  [FAIL] 市级整体: {exc}")

    # 高新通
    if GaoXinTongSpider:
        print("\n--- 高新通 ---")
        try:
            spider = GaoXinTongSpider()
            result = spider.run()
            _accumulate([result])
        except Exception as exc:  # noqa: BLE001
            print(f"  [FAIL] 高新通: {exc}")
            summary['failed'] += 1

    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 爬取完成")
    print(f"  成功 {summary['success']} 个，失败 {summary['failed']} 个，"
          f"采集 {summary['total']} 条，新入库 {summary['inserted']} 条")


def _fetch_articles():
    """从数据库读取所有文章"""
    db = get_db()
    try:
        cursor = db.execute('''
            SELECT title, url, source_name, publish_date, summary,
                   matched_keywords, crawl_time, level
            FROM articles
            ORDER BY crawl_time DESC, id DESC
        ''')
        rows = cursor.fetchall()
    finally:
        db.close()
    return rows


HEADERS = ['标题', '链接', '来源', '发布日期', '摘要', '匹配关键词', '爬取时间', '级别']


def _row_values(article):
    """统一从 sqlite3.Row 中按列名取值"""
    return [
        article['title'],
        article['url'],
        article['source_name'],
        article['publish_date'],
        article['summary'],
        article['matched_keywords'],
        article['crawl_time'],
        article['level'],
    ]


def _export_csv(articles, output_dir, timestamp):
    csv_path = os.path.join(output_dir, f'政策信息_{timestamp}.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for article in articles:
            writer.writerow(_row_values(article))
    print(f"  CSV 已导出: {csv_path}")
    return csv_path


def _export_excel(articles, output_dir, timestamp):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [警告] openpyxl 未安装，跳过 Excel 导出")
        return None

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "政策信息"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1890FF", end_color="1890FF",
                                  fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        def _write_sheet(sheet, rows):
            # 写表头
            for col, header in enumerate(HEADERS, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写数据
            for row_idx, article in enumerate(rows, 2):
                values = _row_values(article)
                for col, value in enumerate(values, 1):
                    sheet.cell(row=row_idx, column=col, value=value)

            # 列宽
            widths = {
                'A': 50, 'B': 40, 'C': 20, 'D': 12,
                'E': 40, 'F': 30, 'G': 20, 'H': 8,
            }
            for col_letter, width in widths.items():
                sheet.column_dimensions[col_letter].width = width

            # 冻结表头
            sheet.freeze_panes = 'A2'

        # 主 Sheet：全部
        _write_sheet(ws, articles)

        # 按级别分 Sheet
        for level_name in ['国家', '省', '市']:
            level_articles = [a for a in articles if a['level'] == level_name]
            if level_articles:
                ws_level = wb.create_sheet(title=f"{level_name}级")
                _write_sheet(ws_level, level_articles)

        xlsx_path = os.path.join(output_dir, f'政策信息_{timestamp}.xlsx')
        wb.save(xlsx_path)
        print(f"  Excel 已导出: {xlsx_path}")
        return xlsx_path
    except Exception as exc:  # noqa: BLE001
        print(f"  [错误] Excel 导出失败: {exc}")
        return None


def export_to_excel_csv():
    """将数据库中的文章导出为 Excel 和 CSV"""
    output_dir = os.path.join(ROOT_DIR, 'output')
    os.makedirs(output_dir, exist_ok=True)

    articles = _fetch_articles()
    if not articles:
        print("  数据库中没有文章数据，跳过导出")
        return

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    _export_csv(articles, output_dir, timestamp)
    _export_excel(articles, output_dir, timestamp)


def main():
    init_db()

    print("=" * 50)
    print("  公共事务部政策信息监控 - 爬取并导出")
    print("=" * 50)
    print()

    run_all_spiders()

    print()
    print("-" * 50)
    print("导出数据...")

    export_to_excel_csv()

    print()
    print("=" * 50)
    print("  完成！请查看 output 目录")
    print("=" * 50)


if __name__ == '__main__':
    main()
