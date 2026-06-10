# -*- coding: utf-8 -*-
"""
公共事务部信息爬取工具 - Flask 应用入口
"""

import csv
import io
import os
import time
import uuid
import threading
import logging
from logging.handlers import TimedRotatingFileHandler
from datetime import datetime, timedelta
from math import ceil

# -------------------------------------------------------------------- #
# 日志配置：控制台 + 文件，按天轮转，保留 30 天
# -------------------------------------------------------------------- #
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

_log_formatter = logging.Formatter(
    '%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

_file_handler = TimedRotatingFileHandler(
    os.path.join(LOG_DIR, 'app.log'),
    when='midnight',
    interval=1,
    backupCount=30,
    encoding='utf-8',
)
_file_handler.setFormatter(_log_formatter)
_file_handler.setLevel(logging.INFO)

_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_log_formatter)
_console_handler.setLevel(logging.INFO)

_root_logger = logging.getLogger()
_root_logger.setLevel(logging.INFO)
# 避免重复加入 handler（debug 重载会二次 import）
if not any(isinstance(h, TimedRotatingFileHandler) for h in _root_logger.handlers):
    _root_logger.addHandler(_file_handler)
if not any(isinstance(h, logging.StreamHandler)
           and not isinstance(h, TimedRotatingFileHandler)
           for h in _root_logger.handlers):
    _root_logger.addHandler(_console_handler)

# 降低三方库日志级别
logging.getLogger('urllib3').setLevel(logging.WARNING)
logging.getLogger('playwright').setLevel(logging.WARNING)
logging.getLogger('werkzeug').setLevel(logging.WARNING)
logging.getLogger('apscheduler').setLevel(logging.WARNING)

from flask import Flask, render_template, request, jsonify, abort, Response, send_file

from config import KEYWORDS, CRAWL_INTERVAL
from database import (
    init_db,
    get_db,
    get_keywords,
    get_keywords_grouped,
    add_keyword,
    add_keywords_batch,
    delete_keyword,
    get_websites,
    get_website,
    add_website,
    update_website,
    delete_website,
)


PER_PAGE = 20

EXPORT_HEADERS = ['标题', '链接', '来源', '发布日期',
                  '匹配关键词', '爬取时间', '级别']


# -------------------------------------------------------------------- #
# 全局爬取进度状态（供轮询使用）
# -------------------------------------------------------------------- #
crawl_progress_lock = threading.Lock()
crawl_progress = {
    'running': False,
    'task_id': None,
    'total': 0,
    'completed': 0,
    'current': '',
    'results': [],          # [{name, status, count, time, error?}]
    'start_time': None,
    'end_time': None,
    'date_from': '',
    'date_to': '',
    'message': '',
    'config': {
        'websites': [],     # 选中的网站名称列表
        'date_from': '',    # 时间范围起始
        'date_to': '',      # 时间范围结束
        'mode': '',         # 'selective' 或 'all'
    },
}

# 爬取控制变量
crawl_control = {
    'cancel': False,   # 取消标志
    'pause': False,    # 暂停标志
}


def _reset_progress(task_id, websites, date_from='', date_to='', mode='selective'):
    with crawl_progress_lock:
        crawl_progress.update({
            'running': True,
            'task_id': task_id,
            'total': len(websites),
            'completed': 0,
            'current': '',
            'results': [],
            'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'end_time': None,
            'date_from': date_from or '',
            'date_to': date_to or '',
            'message': '',
            'config': {
                'websites': [w.get('name', '') for w in websites],
                'date_from': date_from or '',
                'date_to': date_to or '',
                'mode': mode or 'selective',
            },
        })


def _progress_snapshot():
    with crawl_progress_lock:
        # 浅拷贝，results 列表复制，避免外部修改
        snap = dict(crawl_progress)
        snap['results'] = list(crawl_progress['results'])
        return snap


def _get_spider_class_map():
    """返回 {网站名: 爬虫类}。

    复用 scheduler._iter_spider_classes 的动态发现逻辑。
    """
    try:
        from scheduler import _iter_spider_classes
    except Exception as exc:  # noqa: BLE001
        print(f'[crawl] 加载爬虫模块失败: {exc}')
        return {}
    mapping = {}
    for cls in _iter_spider_classes():
        n = (getattr(cls, 'name', None)
             or getattr(cls, 'website_name', None)
             or cls.__name__)
        mapping[n] = cls
    return mapping


def _resolve_target_websites(website_keys):
    """将前端传入的网站 key（名称）解析为 {name, class} 列表。

    website_keys 为空 / None 表示全部。
    """
    spider_map = _get_spider_class_map()
    if not spider_map:
        return []
    if not website_keys:
        names = list(spider_map.keys())
    else:
        names = [n for n in website_keys if n in spider_map]
    return [{'name': n, 'class': spider_map[n]} for n in names]


def _run_crawl_task(task_id, websites, date_from=None, date_to=None):
    """后台线程中逐个调起爬虫，同步更新全局进度状态。"""
    global crawl_control
    from database import update_website_crawl_status

    crawl_control['cancel'] = False
    crawl_control['pause'] = False

    # _reset_progress 已在 _start_crawl_task 中调用，此处不再重复

    cancelled = False
    for spider_info in websites:
        # 检查取消
        if crawl_control['cancel']:
            cancelled = True
            with crawl_progress_lock:
                crawl_progress['current'] = '已取消'
            break

        # 检查暂停 - 循环等待直到恢复
        while crawl_control['pause']:
            if crawl_control['cancel']:
                break
            time.sleep(1)
            with crawl_progress_lock:
                crawl_progress['current'] = f"已暂停 (已完成 {crawl_progress['completed']}/{crawl_progress['total']})"

        if crawl_control['cancel']:
            cancelled = True
            with crawl_progress_lock:
                crawl_progress['current'] = '已取消'
            break

        with crawl_progress_lock:
            crawl_progress['current'] = spider_info['name']
        start = time.time()
        spider = None
        try:
            spider = spider_info['class']()
            if date_from:
                spider.date_from = date_from
            if date_to:
                spider.date_to = date_to
            result = spider.run() or {}
            elapsed = time.time() - start
            entry = {
                'name': spider_info['name'],
                'status': result.get('status') or 'success',
                'count': result.get('inserted', result.get('total', 0)) or 0,
                'total': result.get('total', 0) or 0,
                'total_crawled': result.get('total_crawled', 0) or 0,
                'time': f'{elapsed:.1f}s',
            }
            if result.get('error'):
                entry['status'] = 'failed'
                entry['error'] = result['error']
            try:
                update_website_crawl_status(
                    spider_info['name'],
                    'success' if entry['status'] == 'success' else 'error',
                )
            except Exception:  # noqa: BLE001
                pass
        except Exception as exc:  # noqa: BLE001
            elapsed = time.time() - start
            entry = {
                'name': spider_info['name'],
                'status': 'failed',
                'count': 0,
                'total': 0,
                'total_crawled': 0,
                'time': f'{elapsed:.1f}s',
                'error': f'{type(exc).__name__}: {exc}',
            }
            try:
                update_website_crawl_status(spider_info['name'], 'error')
            except Exception:  # noqa: BLE001
                pass
        finally:
            try:
                if spider is not None:
                    spider.close()
            except Exception:  # noqa: BLE001
                pass

        with crawl_progress_lock:
            crawl_progress['results'].append(entry)
            crawl_progress['completed'] += 1

    with crawl_progress_lock:
        crawl_progress['running'] = False
        crawl_progress['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if cancelled:
            crawl_progress['current'] = f'已取消 (完成 {crawl_progress["completed"]}/{crawl_progress["total"]})'
            crawl_progress['message'] = f'爬取已取消：完成 {crawl_progress["completed"]}/{crawl_progress["total"]} 个网站'
        else:
            crawl_progress['current'] = ''
            success_n = sum(
                1 for r in crawl_progress['results'] if r.get('status') == 'success'
            )
            total_inserted = sum(
                r.get('count', 0) or 0 for r in crawl_progress['results']
            )
            crawl_progress['message'] = (
                f'爬取完成：成功 {success_n}/{len(crawl_progress["results"])} '
                f'个网站，新增入库 {total_inserted} 条'
            )


def _start_crawl_task(websites, date_from=None, date_to=None, mode='selective'):
    """启动后台爬取任务。运行中会被拒绝。返回 (ok, payload)。"""
    if not websites:
        return False, {'status': 'error', 'message': '未找到匹配的爬虫类'}
    with crawl_progress_lock:
        if crawl_progress['running']:
            return False, {
                'status': 'busy',
                'message': '已有爬取任务正在运行，请稍后重试',
                'task_id': crawl_progress['task_id'],
            }
    task_id = uuid.uuid4().hex[:12]
    # 在启动线程前就设置 running=True，避免前端首次轮询时竞态条件
    _reset_progress(task_id, websites, date_from, date_to, mode)
    thread = threading.Thread(
        target=_run_crawl_task,
        args=(task_id, websites, date_from, date_to),
        daemon=True,
    )
    thread.start()
    return True, {
        'status': 'started',
        'task_id': task_id,
        'total': len(websites),
        'message': f'已启动爬取任务，共 {len(websites)} 个网站',
    }



def _export_row_values(article):
    """从查询结果 row（dict）中取出导出所需的字段值"""
    return [
        article.get('title') or '',
        article.get('url') or '',
        article.get('source_name') or '',
        article.get('publish_date') or '',
        article.get('matched_keywords') or '',
        article.get('crawl_time') or '',
        article.get('level') or '',
    ]


def _build_excel_bytes(rows):
    """根据 rows 生成 Excel 文件的二进制内容。

    未安装 openpyxl 时会抛 ImportError。主 Sheet 为全部文章，
    并按 国家/省/市 级别增加分 Sheet。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '政策信息'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1890FF', end_color='1890FF',
                              fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    widths = {'A': 50, 'B': 40, 'C': 20, 'D': 12,
              'E': 30, 'F': 20, 'G': 8}

    def _write_sheet(sheet, items):
        for col, header in enumerate(EXPORT_HEADERS, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_idx, article in enumerate(items, 2):
            for col, value in enumerate(_export_row_values(article), 1):
                sheet.cell(row=row_idx, column=col, value=value)
        for col_letter, width in widths.items():
            sheet.column_dimensions[col_letter].width = width
        sheet.freeze_panes = 'A2'

    _write_sheet(ws, rows)

    for level_name in ['国家', '省', '市']:
        level_rows = [r for r in rows if (r.get('level') or '') == level_name]
        if level_rows:
            ws_level = wb.create_sheet(title=f'{level_name}级')
            _write_sheet(ws_level, level_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def create_app():
    """创建并初始化 Flask 应用"""
    app = Flask(__name__)

    # 确保 data/ 目录存在
    base_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, 'data')
    os.makedirs(data_dir, exist_ok=True)

    # 初始化数据库（首次启动会从 config.py 导入默认关键词与网站）
    init_db()

    # ---------------------------------------------------------------- #
    # 路由
    # ---------------------------------------------------------------- #

    @app.route('/', methods=['GET'])
    def index():
        """仪表盘首页"""
        stats = _get_stats()
        latest_articles = _query_articles(limit=10)['items']
        site_status = _get_site_status_list()
        all_sites = get_websites()
        # 按级别分组，供爬取面板多选使用
        sites_grouped = {'国家': [], '省': [], '市': [], '其他': []}
        for site in all_sites:
            lvl = site.get('level') or '其他'
            if lvl not in sites_grouped:
                sites_grouped[lvl] = []
            sites_grouped[lvl].append(site)
        return render_template(
            'index.html',
            stats=stats,
            latest_articles=latest_articles,
            site_status=site_status,
            websites=all_sites,
            sites_grouped=sites_grouped,
            crawl_interval=CRAWL_INTERVAL,
        )

    @app.route('/results', methods=['GET'])
    def results():
        """爬取结果列表页"""
        filters = _read_filters(request.args)
        page = max(1, int(request.args.get('page', 1) or 1))
        data = _query_articles(filters=filters, page=page, per_page=PER_PAGE)
        return render_template(
            'results.html',
            articles=data['items'],
            total=data['total'],
            page=page,
            per_page=PER_PAGE,
            total_pages=data['total_pages'],
            filters=filters,
            websites=get_websites(),
        )

    @app.route('/api/articles', methods=['GET'])
    def api_articles():
        """文章列表 JSON API"""
        filters = _read_filters(request.args)
        page = max(1, int(request.args.get('page', 1) or 1))
        per_page = int(request.args.get('per_page', PER_PAGE) or PER_PAGE)
        data = _query_articles(filters=filters, page=page, per_page=per_page)
        return jsonify({
            'status': 'ok',
            'page': page,
            'per_page': per_page,
            'total': data['total'],
            'total_pages': data['total_pages'],
            'items': data['items'],
        })

    @app.route('/api/crawl', methods=['POST'])
    def api_crawl():
        """手动触发爬取（后台线程）。并接入全局进度状态。"""
        payload = request.get_json(silent=True) or {}
        website_name = (payload.get('website_name')
                        or request.form.get('website_name'))
        website_keys = [website_name] if website_name else None
        targets = _resolve_target_websites(website_keys)
        mode = 'selective' if website_name else 'all'
        ok, info = _start_crawl_task(targets, mode=mode)
        if ok:
            target_label = website_name if website_name else '全部网站'
            info['target'] = target_label
            info['message'] = f'已启动爬取任务：{target_label}'
            return jsonify(info)
        return jsonify(info), 409

    @app.route('/api/crawl/selective', methods=['POST'])
    def api_crawl_selective():
        """选择性爬取：指定网站 + 可选日期范围。"""
        payload = request.get_json(silent=True) or {}
        websites = payload.get('websites') or []
        date_from = (payload.get('date_from') or '').strip() or None
        date_to = (payload.get('date_to') or '').strip() or None
        targets = _resolve_target_websites(websites)
        if not targets:
            return jsonify({
                'status': 'error',
                'message': '请选择至少一个有效的网站',
            }), 400
        ok, info = _start_crawl_task(targets, date_from, date_to, mode='selective')
        return jsonify(info), (200 if ok else 409)

    @app.route('/api/crawl/all', methods=['POST'])
    def api_crawl_all():
        """爬取全部网站（可选日期范围）。"""
        payload = request.get_json(silent=True) or {}
        date_from = (payload.get('date_from') or '').strip() or None
        date_to = (payload.get('date_to') or '').strip() or None
        targets = _resolve_target_websites(None)
        ok, info = _start_crawl_task(targets, date_from, date_to, mode='all')
        return jsonify(info), (200 if ok else 409)

    @app.route('/api/crawl/cancel', methods=['POST'])
    def api_crawl_cancel():
        """取消爬取"""
        global crawl_control
        if not crawl_progress.get('running'):
            return jsonify({'error': '当前没有正在执行的爬取任务'}), 400
        crawl_control['cancel'] = True
        return jsonify({'success': True, 'message': '正在取消...'})

    @app.route('/api/crawl/pause', methods=['POST'])
    def api_crawl_pause():
        """暂停/恢复爬取"""
        global crawl_control
        if not crawl_progress.get('running'):
            return jsonify({'error': '当前没有正在执行的爬取任务'}), 400
        crawl_control['pause'] = not crawl_control['pause']
        return jsonify({
            'success': True,
            'paused': crawl_control['pause'],
            'message': '已暂停' if crawl_control['pause'] else '已恢复',
        })

    @app.route('/api/crawl/progress', methods=['GET'])
    def api_crawl_progress():
        """返回当前爬取任务进度快照"""
        data = _progress_snapshot()
        data['paused'] = crawl_control.get('pause', False)
        data['cancelled'] = crawl_control.get('cancel', False)
        return jsonify(data)

    @app.route('/api/status', methods=['GET'])
    def api_status():
        """各网站爬取状态 JSON"""
        return jsonify({
            'status': 'ok',
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'sites': _get_site_status_list(),
        })

    # ---------------- 关键词管理 ---------------- #

    @app.route('/keywords', methods=['GET'])
    def keywords_page():
        """关键词管理页面"""
        all_rows = get_keywords()
        grouped = {'general': [], 'ivd': []}
        for row in all_rows:
            grouped.setdefault(row['category'], []).append(row)
        category_labels = {'general': '通用关键词', 'ivd': 'IVD 专业'}
        return render_template(
            'keywords.html',
            grouped=grouped,
            category_labels=category_labels,
            total=len(all_rows),
        )

    @app.route('/api/keywords', methods=['GET'])
    def api_keywords_list():
        """返回所有关键词"""
        return jsonify({
            'status': 'ok',
            'items': get_keywords(),
        })

    @app.route('/api/keywords', methods=['POST'])
    def api_keywords_add():
        """添加关键词，支持批量（逗号分隔）"""
        payload = request.get_json(silent=True) or {}
        keyword = (payload.get('keyword') or request.form.get('keyword') or '').strip()
        category = (payload.get('category') or request.form.get('category')
                    or 'general').strip()
        if category not in ('general', 'ivd'):
            category = 'general'
        if not keyword:
            return jsonify({'status': 'error', 'message': '关键词不能为空'}), 400

        # 批量识别：包含逗号则按批量处理
        if ',' in keyword or '，' in keyword:
            result = add_keywords_batch(keyword, category)
            return jsonify({
                'status': 'ok',
                'message': f"成功添加 {len(result['added'])} 个，跳过 {len(result['skipped'])} 个",
                'added': result['added'],
                'skipped': result['skipped'],
            })

        ok, message = add_keyword(keyword, category)
        return jsonify({
            'status': 'ok' if ok else 'error',
            'message': message,
        }), (200 if ok else 400)

    @app.route('/api/keywords/<int:keyword_id>', methods=['DELETE'])
    def api_keywords_delete(keyword_id):
        """删除关键词"""
        ok = delete_keyword(keyword_id)
        return jsonify({
            'status': 'ok' if ok else 'error',
            'message': '删除成功' if ok else '关键词不存在',
        }), (200 if ok else 404)

    # ---------------- 网址管理 ---------------- #

    @app.route('/websites', methods=['GET'])
    def websites_page():
        """网址管理页面"""
        sites = get_websites()
        # 同时带上状态信息
        status_map = {s['name']: s for s in _get_site_status_list()}
        for site in sites:
            info = status_map.get(site['name'], {})
            site['runtime_status'] = info.get('status', 'idle')
            site['last_crawl_time'] = (site.get('last_crawl_time')
                                       or info.get('last_crawl_time') or '')
        return render_template('websites.html', websites=sites)

    @app.route('/api/websites', methods=['GET'])
    def api_websites_list():
        return jsonify({'status': 'ok', 'items': get_websites()})

    @app.route('/api/websites', methods=['POST'])
    def api_websites_add():
        payload = request.get_json(silent=True) or request.form.to_dict()
        name = payload.get('name')
        url = payload.get('url')
        level = payload.get('level') or '国家'
        buttons = payload.get('buttons') or ''
        ok, message, new_id = add_website(name, url, level, buttons)
        return jsonify({
            'status': 'ok' if ok else 'error',
            'message': message,
            'id': new_id,
        }), (200 if ok else 400)

    @app.route('/api/websites/<int:website_id>', methods=['PUT'])
    def api_websites_update(website_id):
        payload = request.get_json(silent=True) or request.form.to_dict()
        ok, message = update_website(
            website_id,
            name=payload.get('name'),
            url=payload.get('url'),
            level=payload.get('level'),
            buttons=payload.get('buttons'),
        )
        return jsonify({
            'status': 'ok' if ok else 'error',
            'message': message,
        }), (200 if ok else 400)

    @app.route('/api/websites/<int:website_id>', methods=['DELETE'])
    def api_websites_delete(website_id):
        ok = delete_website(website_id)
        return jsonify({
            'status': 'ok' if ok else 'error',
            'message': '删除成功' if ok else '网站不存在',
        }), (200 if ok else 404)

    # ---------------- 数据管理（删除） ---------------- #

    @app.route('/api/articles/<int:article_id>', methods=['DELETE'])
    def api_delete_article(article_id):
        """删除单条文章"""
        conn = get_db()
        cursor = conn.execute('DELETE FROM articles WHERE id = ?', (article_id,))
        conn.commit()
        rowcount = cursor.rowcount
        conn.close()
        if rowcount > 0:
            return jsonify({'success': True, 'message': '已删除'})
        else:
            return jsonify({'error': '文章不存在'}), 404

    @app.route('/api/articles/delete', methods=['POST'])
    def api_delete_articles():
        """删除文章"""
        data = request.json or {}
        mode = data.get('mode')  # 'filtered', 'all', 'by_date', 'no_keywords'

        conn = get_db()

        if mode == 'all':
            # 清除全部
            cursor = conn.execute('DELETE FROM articles')

        elif mode == 'filtered':
            # 清除当前筛选条件的
            source = data.get('source', '')
            date_from = data.get('date_from', '')
            date_to = data.get('date_to', '')
            keyword = data.get('keyword', '')

            query = 'DELETE FROM articles WHERE 1=1'
            params = []
            if source:
                query += ' AND source_name = ?'
                params.append(source)
            if date_from:
                query += ' AND publish_date >= ?'
                params.append(date_from)
            if date_to:
                query += ' AND publish_date <= ?'
                params.append(date_to)
            if keyword:
                query += ' AND (title LIKE ? OR matched_keywords LIKE ?)'
                params.extend([f'%{keyword}%', f'%{keyword}%'])
            cursor = conn.execute(query, params)

        elif mode == 'by_date':
            # 清除指定日期范围
            date_from = data.get('date_from', '')
            date_to = data.get('date_to', '')
            query = 'DELETE FROM articles WHERE 1=1'
            params = []
            if date_from:
                query += ' AND publish_date >= ?'
                params.append(date_from)
            if date_to:
                query += ' AND publish_date <= ?'
                params.append(date_to)
            cursor = conn.execute(query, params)

        elif mode == 'no_keywords':
            # 清除没有匹配关键词的
            cursor = conn.execute(
                "DELETE FROM articles WHERE matched_keywords IS NULL "
                "OR matched_keywords = '' OR matched_keywords = '[]'"
            )

        else:
            conn.close()
            return jsonify({'error': '未知操作'}), 400

        deleted = cursor.rowcount
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'deleted': deleted})

    # ---------------- 数据导出 ---------------- #

    @app.route('/api/export/csv', methods=['GET'])
    def api_export_csv():
        """按筛选条件导出 CSV，浏览器直接下载"""
        filters = _read_filters(request.args)
        rows = _query_articles(filters=filters, limit=100000)['items']

        buffer = io.StringIO()
        # utf-8-sig：写入 BOM，使 Excel 打开不乱码
        buffer.write('\ufeff')
        writer = csv.writer(buffer)
        writer.writerow(EXPORT_HEADERS)
        for row in rows:
            writer.writerow(_export_row_values(row))

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'政策信息_{timestamp}.csv'
        ascii_fallback = f'policy_{timestamp}.csv'
        # RFC 5987：兼容中文文件名
        from urllib.parse import quote
        disposition = (
            f"attachment; filename={ascii_fallback}; "
            f"filename*=UTF-8''{quote(filename)}"
        )
        return Response(
            buffer.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': disposition},
        )

    @app.route('/api/export/excel', methods=['GET'])
    def api_export_excel():
        """按筛选条件导出 Excel，浏览器直接下载"""
        filters = _read_filters(request.args)
        rows = _query_articles(filters=filters, limit=100000)['items']

        try:
            xlsx_bytes = _build_excel_bytes(rows)
        except ImportError:
            return jsonify({
                'status': 'error',
                'message': '未安装 openpyxl，无法导出 Excel',
            }), 500

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'政策信息_{timestamp}.xlsx'
        return send_file(
            io.BytesIO(xlsx_bytes),
            mimetype=('application/vnd.openxmlformats-officedocument'
                      '.spreadsheetml.sheet'),
            as_attachment=True,
            download_name=filename,
        )

    return app


# -------------------------------------------------------------------- #
# 辅助函数
# -------------------------------------------------------------------- #

def _seed_websites():
    """预留接口，默认网站现在由 init_db() 负责导入"""
    pass


def _read_filters(args):
    """从 query string 中读取筛选条件"""
    return {
        'source': (args.get('source') or '').strip(),
        'level': (args.get('level') or '').strip(),
        'date_from': (args.get('date_from') or '').strip(),
        'date_to': (args.get('date_to') or '').strip(),
        'keyword': (args.get('keyword') or '').strip(),
    }


def _query_articles(filters=None, page=1, per_page=PER_PAGE, limit=None):  # noqa: C901
    """统一的文章查询函数"""
    filters = filters or {}
    where_parts = []
    params = []

    if filters.get('source'):
        where_parts.append('source_name = ?')
        params.append(filters['source'])
    if filters.get('level'):
        where_parts.append('level = ?')
        params.append(filters['level'])
    if filters.get('date_from'):
        where_parts.append('publish_date >= ?')
        params.append(filters['date_from'])
    if filters.get('date_to'):
        where_parts.append('publish_date <= ?')
        params.append(filters['date_to'])
    if filters.get('keyword'):
        where_parts.append('(title LIKE ? OR matched_keywords LIKE ?)')
        kw = f"%{filters['keyword']}%"
        params.extend([kw, kw])

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    conn = get_db()
    cursor = conn.cursor()

    # 总数
    cursor.execute(f'SELECT COUNT(*) AS c FROM articles {where_sql}', params)
    total = cursor.fetchone()['c']

    # 列表
    if limit is not None:
        sql = (f'SELECT * FROM articles {where_sql} '
               f'ORDER BY COALESCE(publish_date, crawl_time) DESC, id DESC LIMIT ?')
        cursor.execute(sql, params + [limit])
    else:
        offset = (page - 1) * per_page
        sql = (f'SELECT * FROM articles {where_sql} '
               f'ORDER BY COALESCE(publish_date, crawl_time) DESC, id DESC '
               f'LIMIT ? OFFSET ?')
        cursor.execute(sql, params + [per_page, offset])

    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    # 处理 matched_keywords 字段：变成数组
    for row in rows:
        raw = row.get('matched_keywords') or ''
        row['matched_keywords_list'] = [
            k.strip() for k in raw.replace('，', ',').split(',') if k.strip()
        ]

    total_pages = max(1, ceil(total / per_page)) if per_page else 1
    return {
        'items': rows,
        'total': total,
        'total_pages': total_pages,
    }


def _get_stats():
    """统计数据"""
    today = datetime.now().strftime('%Y-%m-%d')
    week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT COUNT(*) AS c FROM articles "
        "WHERE substr(COALESCE(crawl_time, publish_date),1,10) = ?",
        (today,),
    )
    today_count = cursor.fetchone()['c']

    cursor.execute(
        "SELECT COUNT(*) AS c FROM articles "
        "WHERE substr(COALESCE(crawl_time, publish_date),1,10) >= ?",
        (week_ago,),
    )
    week_count = cursor.fetchone()['c']

    cursor.execute('SELECT COUNT(*) AS c FROM articles')
    total_count = cursor.fetchone()['c']

    conn.close()

    return {
        'today': today_count,
        'week': week_count,
        'total': total_count,
        'sites': len(get_websites()),
    }


def _get_site_status_list():
    """各网站最新爬取状态"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT name, level, url, status, last_crawl_time FROM websites')
    db_rows = {r['name']: dict(r) for r in cursor.fetchall()}

    # 取最近一条爬取日志的状态
    cursor.execute(
        'SELECT website_name, status, end_time, error_message FROM crawl_logs '
        'WHERE id IN (SELECT MAX(id) FROM crawl_logs GROUP BY website_name)'
    )
    log_rows = {r['website_name']: dict(r) for r in cursor.fetchall()}
    conn.close()

    result = []
    for site in get_websites():
        name = site['name']
        db_info = db_rows.get(name, {})
        log_info = log_rows.get(name, {})
        last_time = (db_info.get('last_crawl_time')
                     or log_info.get('end_time')
                     or '')
        log_status = log_info.get('status')
        if log_status == 'success':
            status = 'success'
        elif log_status in ('error', 'failed'):
            status = 'error'
        elif last_time:
            status = 'success'
        else:
            status = 'idle'
        result.append({
            'name': name,
            'level': site['level'],
            'url': site['url'],
            'last_crawl_time': last_time,
            'status': status,
            'error_message': log_info.get('error_message') or '',
        })
    return result


def _run_crawl_async(website_name=None):
    """后台线程中执行爬取，调用 scheduler 中的统一入口"""
    try:
        from scheduler import run_crawl_job
        run_crawl_job(website_name)
    except Exception as exc:  # noqa: BLE001
        print(f'[crawl] 爬取任务执行异常: {exc}')


app = create_app()


# 启动定时任务
def _start_background_scheduler():
    """启动后台定时调度器。start_scheduler 内部有防重复保护。"""
    try:
        from scheduler import start_scheduler
        start_scheduler(app)
    except Exception as exc:  # noqa: BLE001
        print(f'[scheduler] 定时任务未启动: {exc}')


if __name__ == '__main__':
    # debug 模式下，仅在 reloader 子进程中启动调度器（避免父子双开）
    # 非 debug 模式下，直接启动
    use_debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    if not use_debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        _start_background_scheduler()
    app.run(debug=use_debug, host='0.0.0.0', port=5000, use_reloader=use_debug)
